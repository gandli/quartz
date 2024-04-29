import openpyxl
import ollama


def generate_markdown_table_from_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    headers = [cell.value for cell in sheet[1]]
    data_count = sheet.max_row - 1
    table = "| " + " | ".join(headers) + " |\n" + "| --- " * len(headers) + "|\n"
    for row in sheet.iter_rows(values_only=True, min_row=2):
        table += "| " + " | ".join(str(cell) for cell in row) + " |\n"
    return sheet.title, data_count, table


title, data_count, table = generate_markdown_table_from_excel("data.xlsx")

print(title, "\n", table)

response = ollama.chat(
    model="llama3",
    messages=[
        {
            "role": "user",
            "content": "你现在是一名数据分析师，你精通各种统计分析方法，懂得如何清洗、处理和解析数据以获得有价值的洞察。你擅长利用数据驱动的方式来解决问题和提升决策效率。请在这个角色下为我分析数据集，先从多维度分析，再提出关键问题和相应的回答。并始终使用中文回应，符合中文使用习惯。",
        },
        {
            "role": "assistant",
            "content": f"请基于 {title} 数据集的 {data_count} 条数据：\n{table}",
        },
    ],
)
print(response["message"]["content"])
